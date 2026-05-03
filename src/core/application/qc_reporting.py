from __future__ import annotations

import hashlib
import json
import time
from pathlib import Path
from typing import Any, Iterable, Mapping

import pandas as pd
from openpyxl import Workbook


def _utc_ts() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_dump(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(dict(payload), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _coerce_row(item: Any) -> dict[str, Any]:
    if isinstance(item, Mapping):
        return dict(item)
    if hasattr(item, "to_dict"):
        maybe = item.to_dict()
        if isinstance(maybe, Mapping):
            return dict(maybe)
    if hasattr(item, "__dict__"):
        return {k: v for k, v in vars(item).items() if not k.startswith("_")}
    raise TypeError(f"Unsupported QC payload row: {type(item)!r}")


def issues_to_frame(issues: Iterable[Any]) -> pd.DataFrame:
    rows = [_coerce_row(item) for item in issues]
    base_cols = [
        "qc_run",
        "data_version",
        "rule_id",
        "domain",
        "dataset",
        "severity",
        "message",
        "remediation",
        "row_id",
        "field",
        "sample_value",
        "check",
    ]
    df = pd.DataFrame(rows)
    for col in base_cols:
        if col not in df.columns:
            df[col] = None
    ordered = [c for c in base_cols if c in df.columns] + [c for c in df.columns if c not in base_cols]
    return df[ordered]


def alerts_to_frame(alerts: Iterable[Any]) -> pd.DataFrame:
    rows = [_coerce_row(item) for item in alerts]
    base_cols = [
        "tenant_id",
        "alert_id",
        "farm_id",
        "alert_date",
        "severity",
        "alert_type",
        "entity_type",
        "entity_id",
        "message",
        "source_rule_id",
        "qc_run",
        "data_version",
    ]
    df = pd.DataFrame(rows)
    for col in base_cols:
        if col not in df.columns:
            df[col] = None
    ordered = [c for c in base_cols if c in df.columns] + [c for c in df.columns if c not in base_cols]
    return df[ordered]


def build_issue_counts(issues_df: pd.DataFrame) -> dict[str, int]:
    if issues_df.empty or "severity" not in issues_df.columns:
        return {}
    counts = issues_df["severity"].fillna("UNKNOWN").astype(str).value_counts().to_dict()
    return {str(k): int(v) for k, v in counts.items()}


def build_row_counts(issues_df: pd.DataFrame) -> dict[str, int]:
    if issues_df.empty:
        return {}
    if "dataset" not in issues_df.columns:
        return {"issues_total": int(len(issues_df))}
    out: dict[str, int] = {}
    for dataset, sub in issues_df.groupby(issues_df["dataset"].fillna(""), dropna=False):
        ds_key = str(dataset or "unknown")
        out[f"{ds_key}.issues"] = int(len(sub))
        if "row_id" in sub.columns:
            non_empty = sub["row_id"].fillna("").astype(str)
            unique_rows = non_empty[non_empty != ""].nunique()
            out[f"{ds_key}.rows"] = int(unique_rows)
    out["issues_total"] = int(len(issues_df))
    return out


def build_bad_rows_frames(issues_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if issues_df.empty:
        empty = pd.DataFrame(columns=["row_id", "reason"])
        return empty, issues_df.copy()

    detailed = issues_df.copy()
    if "row_id" not in detailed.columns:
        detailed["row_id"] = ""
    if "check" not in detailed.columns:
        detailed["check"] = detailed.get("rule_id", "")

    def _build_reason(row: pd.Series) -> str:
        sev = str(row.get("severity") or "")
        check = str(row.get("check") or row.get("rule_id") or "")
        msg = str(row.get("message") or "")
        parts = [p for p in [sev, check, msg] if p]
        return ": ".join(parts)

    detailed["reason"] = detailed.apply(_build_reason, axis=1)
    collapsed = (
        detailed.assign(row_id=detailed["row_id"].fillna("").astype(str))
        .groupby("row_id", dropna=False, as_index=False)["reason"]
        .agg(lambda xs: " | ".join([str(x) for x in xs if str(x).strip()]))
    )
    return collapsed[["row_id", "reason"]], detailed


def write_qc_report_xlsx(path: Path, *, summary: Mapping[str, Any], issues_df: pd.DataFrame, alerts_df: pd.DataFrame | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    summary_pairs = [
        ("schema", summary.get("schema")),
        ("created_at_utc", summary.get("created_at_utc") or summary.get("generated_at")),
        ("data_version", summary.get("data_version")),
        ("qc_run", summary.get("qc_run")),
        ("qc_status", summary.get("qc_status")),
        ("config_version", summary.get("config_version")),
        ("config_path", summary.get("config_path")),
        ("rules_sha256", summary.get("rules_sha256")),
    ]
    for key, value in summary_pairs:
        ws.append([key, "" if value is None else value])

    ws.append([])
    ws.append(["issue_count", "value"])
    for key, value in sorted(dict(summary.get("issue_counts") or {}).items()):
        ws.append([key, value])

    ws.append([])
    ws.append(["row_count", "value"])
    for key, value in sorted(dict(summary.get("row_counts") or {}).items()):
        ws.append([key, value])

    ws.append([])
    ws.append(["metric", "value"])
    for key, value in sorted(dict(summary.get("metrics") or {}).items()):
        ws.append([key, value])

    ws_issues = wb.create_sheet("Issues")
    if issues_df.empty:
        ws_issues.append(["message"])
        ws_issues.append(["No QC issues"])
    else:
        ws_issues.append(list(issues_df.columns))
        for row in issues_df.fillna("").itertuples(index=False, name=None):
            ws_issues.append(list(row))

    ws_bad = wb.create_sheet("BadRows")
    bad_rows_df, _ = build_bad_rows_frames(issues_df)
    ws_bad.append(list(bad_rows_df.columns))
    for row in bad_rows_df.fillna("").itertuples(index=False, name=None):
        ws_bad.append(list(row))

    if alerts_df is not None:
        ws_alerts = wb.create_sheet("Alerts")
        if alerts_df.empty:
            ws_alerts.append(["message"])
            ws_alerts.append(["No auto alerts"])
        else:
            ws_alerts.append(list(alerts_df.columns))
            for row in alerts_df.fillna("").itertuples(index=False, name=None):
                ws_alerts.append(list(row))

    wb.save(path)


def write_qc_output_bundle(
    *,
    out_dir: Path,
    summary: Mapping[str, Any],
    issues: Iterable[Any],
    alerts: Iterable[Any] | None = None,
    include_alerts_csv: bool = False,
    include_bad_rows_detailed: bool = False,
    manifest_type: str = "qc",
) -> dict[str, Any]:
    out_dir.mkdir(parents=True, exist_ok=True)

    issues_df = issues_to_frame(issues)
    alerts_df = alerts_to_frame(alerts or [])
    bad_rows_df, detailed_bad_rows_df = build_bad_rows_frames(issues_df)

    outputs: dict[str, str] = {}

    issues_path = out_dir / "qc_issues.csv"
    issues_df.to_csv(issues_path, index=False)
    outputs["qc_issues_csv"] = str(issues_path)

    bad_rows_path = out_dir / "bad_rows.csv"
    bad_rows_df.to_csv(bad_rows_path, index=False)
    outputs["bad_rows_csv"] = str(bad_rows_path)

    if include_bad_rows_detailed:
        bad_rows_detailed_path = out_dir / "bad_rows_detailed.csv"
        detailed_bad_rows_df.to_csv(bad_rows_detailed_path, index=False)
        outputs["bad_rows_detailed_csv"] = str(bad_rows_detailed_path)

    if include_alerts_csv:
        alerts_path = out_dir / "alerts_auto.csv"
        alerts_df.to_csv(alerts_path, index=False)
        outputs["alerts_auto_csv"] = str(alerts_path)

    summary_payload = dict(summary)
    summary_payload.setdefault("created_at_utc", summary_payload.get("generated_at") or _utc_ts())
    if "issue_counts" in summary_payload and summary_payload.get("issue_counts") is not None:
        summary_payload["issue_counts"] = dict(summary_payload.get("issue_counts") or {})
    else:
        summary_payload["issue_counts"] = build_issue_counts(issues_df)
    if "row_counts" in summary_payload and summary_payload.get("row_counts") is not None:
        summary_payload["row_counts"] = dict(summary_payload.get("row_counts") or {})
    else:
        summary_payload["row_counts"] = build_row_counts(issues_df)
    summary_payload["outputs"] = {**dict(summary_payload.get("outputs") or {}), **outputs}

    report_path = out_dir / "qc_report.xlsx"
    write_qc_report_xlsx(report_path, summary=summary_payload, issues_df=issues_df, alerts_df=alerts_df if include_alerts_csv else None)
    outputs["qc_report_xlsx"] = str(report_path)
    summary_payload["outputs"] = {**dict(summary_payload.get("outputs") or {}), "qc_report_xlsx": str(report_path)}

    summary_path = out_dir / "qc_summary.json"
    _json_dump(summary_path, summary_payload)
    outputs["qc_summary_json"] = str(summary_path)
    summary_payload["outputs"] = {**dict(summary_payload.get("outputs") or {}), "qc_summary_json": str(summary_path)}

    manifest_path = out_dir / "manifest.json"
    manifest_paths = [issues_path, bad_rows_path, report_path, summary_path]
    if include_alerts_csv:
        manifest_paths.append(out_dir / "alerts_auto.csv")
    if include_bad_rows_detailed:
        manifest_paths.append(out_dir / "bad_rows_detailed.csv")
    manifest = {
        "type": manifest_type,
        "data_version": summary_payload.get("data_version"),
        "run_id": summary_payload.get("qc_run") or summary_payload.get("run_id"),
        "generated_at": summary_payload.get("created_at_utc") or summary_payload.get("generated_at") or _utc_ts(),
        "config_version": summary_payload.get("config_version"),
        "checksums": {path.name: _sha256_file(path) for path in manifest_paths if path.exists()},
    }
    _json_dump(manifest_path, manifest)
    outputs["manifest_json"] = str(manifest_path)
    summary_payload["outputs"] = {**dict(summary_payload.get("outputs") or {}), "manifest_json": str(manifest_path)}

    _json_dump(summary_path, summary_payload)

    return summary_payload
