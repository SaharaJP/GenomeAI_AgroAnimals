"""Data upload boundary: TYPE_REGISTRY, template generation, parse, validate, commit."""
from __future__ import annotations

import csv
import io
import logging
import secrets
import time
import uuid as _uuid
from dataclasses import dataclass, field
from typing import Any, Optional

from packages.contracts.api_boundary_v1 import (
    UploadColumnSpec,
    UploadCommitResponse,
    UploadPreviewResponse,
    UploadRowError,
    UploadTypeMeta,
)

logger = logging.getLogger("genomeai.web_cabinet.uploads_v1")


@dataclass
class UploadType:
    type_id: str
    label: str
    target_table: str
    columns: list[UploadColumnSpec]
    unique_key: list[str]
    sample_row: dict[str, Any]
    instructions: str = ''
    server_id_column: Optional[str] = None
    auto_fields: dict[str, str] = field(default_factory=dict)


TYPE_REGISTRY: dict[str, UploadType] = {
    'milkings': UploadType(
        type_id='milkings',
        label='Надои',
        target_table='dm_milkings_daily',
        columns=[
            UploadColumnSpec(name='animal_id', required=True, kind='str',
                             description='ID животного (должно существовать)', fk_table='dm_animals'),
            UploadColumnSpec(name='date', required=True, kind='date',
                             description='Дата надоя в формате YYYY-MM-DD'),
            UploadColumnSpec(name='milk_kg', required=True, kind='float',
                             description='Надой в кг (0-80)', min_val=0, max_val=80),
            UploadColumnSpec(name='scc_cells_ml', required=False, kind='int',
                             description='SCC, cells/ml (0-5000000)', min_val=0, max_val=5_000_000),
            UploadColumnSpec(name='fat_pct', required=False, kind='float',
                             description='Жирность %, 0-10', min_val=0, max_val=10),
            UploadColumnSpec(name='protein_pct', required=False, kind='float',
                             description='Белок %, 0-10', min_val=0, max_val=10),
        ],
        unique_key=['tenant_id', 'animal_id', 'date'],
        sample_row={'animal_id': '1234', 'date': '2026-05-01', 'milk_kg': 28.5,
                    'scc_cells_ml': 180000, 'fat_pct': 3.8, 'protein_pct': 3.2},
        instructions='Каждая строка — один день надоя одной коровы.',
        server_id_column='record_id',
        auto_fields={'tenant_id': 'tenant_id'},
    ),
    'health_events': UploadType(
        type_id='health_events',
        label='События здоровья',
        target_table='dm_health_events',
        columns=[
            UploadColumnSpec(name='animal_id', required=True, kind='str',
                             description='ID животного', fk_table='dm_animals'),
            UploadColumnSpec(name='event_date', required=True, kind='date',
                             description='Дата события'),
            UploadColumnSpec(name='event_type', required=True, kind='str',
                             description='Тип: mastitis, lameness, ketosis, vet_visit, ...'),
            UploadColumnSpec(name='severity', required=False, kind='str',
                             description='Опционально: low, medium, high'),
            UploadColumnSpec(name='notes', required=False, kind='str',
                             description='Комментарий'),
        ],
        unique_key=['tenant_id', 'animal_id', 'event_date', 'event_type'],
        sample_row={'animal_id': '1234', 'event_date': '2026-05-01',
                    'event_type': 'mastitis', 'severity': 'medium',
                    'notes': 'Левая передняя четверть'},
        instructions='Регистрация событий здоровья по конкретным животным.',
        server_id_column='event_id',
        auto_fields={'tenant_id': 'tenant_id'},
    ),
    'animals': UploadType(
        type_id='animals',
        label='Животные',
        target_table='dm_animals',
        columns=[
            UploadColumnSpec(name='animal_id', required=True, kind='str',
                             description='Уникальный ID нового животного'),
            UploadColumnSpec(name='external_id', required=False, kind='str',
                             description='Внешний номер бирки'),
            UploadColumnSpec(name='birth_date', required=False, kind='date',
                             description='Дата рождения'),
            UploadColumnSpec(name='sex', required=False, kind='str',
                             description='cow | heifer | bull | calf'),
            UploadColumnSpec(name='breed', required=False, kind='str',
                             description='Порода'),
            UploadColumnSpec(name='status', required=False, kind='str',
                             description='active | dry | culled | sold'),
        ],
        unique_key=['tenant_id', 'animal_id'],
        sample_row={'animal_id': '5001', 'external_id': 'EAR-5001',
                    'birth_date': '2024-03-15', 'sex': 'heifer',
                    'breed': 'Holstein', 'status': 'active'},
        instructions='Регистрация новых животных. ID должен быть уникальным в пределах хозяйства.',
        auto_fields={'tenant_id': 'tenant_id', 'farm_id': 'farm_id'},
    ),
    'feed_rations': UploadType(
        type_id='feed_rations',
        label='Рационы',
        target_table='dm_feed_rations',
        columns=[
            UploadColumnSpec(name='ration_id', required=True, kind='str',
                             description='Уникальный ID рациона'),
            UploadColumnSpec(name='ration_name', required=True, kind='str',
                             description='Название рациона'),
            UploadColumnSpec(name='effective_from', required=True, kind='date',
                             description='Дата начала действия'),
            UploadColumnSpec(name='effective_to', required=False, kind='date',
                             description='Дата окончания (опционально)'),
            UploadColumnSpec(name='dm_pct', required=False, kind='float',
                             description='Сухое вещество %, 0-100', min_val=0, max_val=100),
        ],
        unique_key=['tenant_id', 'ration_id'],
        sample_row={'ration_id': 'R-001', 'ration_name': 'Высокопродуктивные TMR',
                    'effective_from': '2026-05-01', 'effective_to': '',
                    'dm_pct': 50.0},
        instructions='Рационы кормления для разных групп животных.',
        auto_fields={'tenant_id': 'tenant_id', 'site_id': 'site_id'},
    ),
}


def list_types() -> list[UploadTypeMeta]:
    return [
        UploadTypeMeta(
            type=t.type_id, label=t.label, target_table=t.target_table,
            instructions=t.instructions, columns=t.columns,
        )
        for t in TYPE_REGISTRY.values()
    ]


def generate_template(type_id: str, fmt: str) -> tuple[bytes, str, str]:
    """Return (body_bytes, content_type, filename)."""
    if type_id not in TYPE_REGISTRY:
        raise ValueError(f'unknown type: {type_id}')
    if fmt not in ('csv', 'xlsx'):
        raise ValueError(f'unknown format: {fmt}')
    spec = TYPE_REGISTRY[type_id]
    headers = [c.name for c in spec.columns]
    sample = [str(spec.sample_row.get(c.name, '')) for c in spec.columns]

    if fmt == 'csv':
        buf = io.StringIO()
        w = csv.writer(buf, lineterminator='\n')
        w.writerow(headers)
        w.writerow(sample)
        body = ('﻿' + buf.getvalue()).encode('utf-8')  # UTF-8 BOM
        return body, 'text/csv; charset=utf-8', f'{type_id}_template.csv'

    # xlsx
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = type_id
    ws.append(headers)
    instructions = []
    for c in spec.columns:
        bits = [c.kind]
        if c.required:
            bits.append('required')
        if c.min_val is not None or c.max_val is not None:
            mn = c.min_val if c.min_val is not None else '-∞'
            mx = c.max_val if c.max_val is not None else '+∞'
            bits.append(f'[{mn}, {mx}]')
        if c.fk_table:
            bits.append(f'FK→{c.fk_table}')
        if c.description:
            bits.append(c.description)
        instructions.append(' • '.join(bits))
    ws.append(instructions)
    ws.append(sample)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal='left')
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return (
        out.getvalue(),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        f'{type_id}_template.xlsx',
    )


from datetime import datetime


@dataclass
class ValidationResult:
    valid_rows: list[dict]
    duplicates: int
    errors: list[UploadRowError]


def parse_file(type_id: str, file_bytes: bytes, filename: str) -> list[dict]:
    if type_id not in TYPE_REGISTRY:
        raise ValueError(f'unknown type: {type_id}')
    name = filename.lower()
    if name.endswith('.xlsx'):
        return _parse_xlsx(type_id, file_bytes)
    if name.endswith('.csv'):
        return _parse_csv(type_id, file_bytes)
    raise ValueError('unsupported_format')


def _parse_csv(type_id: str, body: bytes) -> list[dict]:
    spec = TYPE_REGISTRY[type_id]
    expected_cols = {c.name for c in spec.columns}
    text = None
    for enc in ('utf-8-sig', 'utf-8', 'cp1251'):
        try:
            text = body.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return []
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict] = []
    for raw in reader:
        if not any(((v or '').strip()) for v in raw.values() if isinstance(v, str)):
            continue
        cleaned: dict[str, Any] = {}
        for k, v in raw.items():
            if k in expected_cols:
                if isinstance(v, str):
                    v = v.strip()
                cleaned[k] = v
        if cleaned:
            rows.append(cleaned)
    return rows


def _parse_xlsx(type_id: str, body: bytes) -> list[dict]:
    from openpyxl import load_workbook
    spec = TYPE_REGISTRY[type_id]
    expected_cols = {c.name for c in spec.columns}
    wb = load_workbook(io.BytesIO(body), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        # Skip xlsx instruction row (template puts ' • '-joined text at row 2)
        if isinstance(row[0], str) and (' • ' in row[0] or row[0].strip() in ('str', 'int', 'float', 'date')):
            continue
        d: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if h in expected_cols and i < len(row):
                v = row[i]
                if isinstance(v, str):
                    v = v.strip()
                if v == '':
                    continue
                if hasattr(v, 'isoformat'):
                    v = v.isoformat()[:10]
                d[h] = v
        if d:
            rows.append(d)
    wb.close()
    return rows


def _coerce(value: Any, kind: str) -> tuple[Any, Optional[str]]:
    """Coerce value to kind. Returns (coerced, error_message_or_None)."""
    if value is None or value == '':
        return None, None
    try:
        if kind == 'int':
            return int(float(str(value))), None
        if kind == 'float':
            return float(str(value)), None
        if kind == 'date':
            s = str(value)
            for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%Y-%m-%dT%H:%M:%S', '%d/%m/%Y'):
                try:
                    return datetime.strptime(s[:19], fmt).date().isoformat(), None
                except ValueError:
                    continue
            return None, 'Неверный формат даты (ожидается YYYY-MM-DD)'
        return str(value).strip(), None
    except (ValueError, TypeError):
        return None, 'Неверный тип'


def validate_rows(type_id: str, rows: list[dict], *, tenant_id: str) -> ValidationResult:
    if type_id not in TYPE_REGISTRY:
        raise ValueError(f'unknown type: {type_id}')
    spec = TYPE_REGISTRY[type_id]
    errors: list[UploadRowError] = []
    coerced_rows: list[dict] = []

    # Pass 1: per-row required + coercion + range
    for i, raw in enumerate(rows, start=1):
        row_errs: list[UploadRowError] = []
        clean: dict[str, Any] = {}
        for col in spec.columns:
            v = raw.get(col.name)
            if col.required and (v is None or v == ''):
                row_errs.append(UploadRowError(row=i, field=col.name, message='Обязательное поле'))
                continue
            coerced, err = _coerce(v, col.kind)
            if err is not None:
                row_errs.append(UploadRowError(row=i, field=col.name, message=f'{err}: {v!r}'))
                continue
            if coerced is None:
                continue
            if col.min_val is not None and isinstance(coerced, (int, float)) and coerced < col.min_val:
                row_errs.append(UploadRowError(row=i, field=col.name,
                    message=f'Значение {coerced} меньше минимума {col.min_val}'))
                continue
            if col.max_val is not None and isinstance(coerced, (int, float)) and coerced > col.max_val:
                row_errs.append(UploadRowError(row=i, field=col.name,
                    message=f'Значение {coerced} больше максимума {col.max_val}'))
                continue
            clean[col.name] = coerced
        if row_errs:
            errors.extend(row_errs)
        else:
            clean['_row_idx'] = i
            coerced_rows.append(clean)

    # Pass 2: FK existence (batch per fk_table)
    fk_cols = [c for c in spec.columns if c.fk_table]
    for col in fk_cols:
        ids = sorted({r[col.name] for r in coerced_rows if col.name in r})
        if not ids:
            continue
        existing = _fetch_fk_existing(col.fk_table, tenant_id, ids)
        kept: list[dict] = []
        for r in coerced_rows:
            if col.name in r and r[col.name] not in existing:
                errors.append(UploadRowError(row=r['_row_idx'], field=col.name,
                    message=f'ID {r[col.name]!r} не существует в {col.fk_table}'))
                continue
            kept.append(r)
        coerced_rows = kept

    # Pass 3: duplicate detection
    duplicates = 0
    if spec.unique_key and coerced_rows:
        existing_keys = _fetch_existing_keys(spec, tenant_id, coerced_rows)
        kept = []
        for r in coerced_rows:
            key = tuple(r.get(k) if k != 'tenant_id' else tenant_id for k in spec.unique_key)
            if key in existing_keys:
                duplicates += 1
                continue
            kept.append(r)
        coerced_rows = kept

    valid_rows = [{k: v for k, v in r.items() if not k.startswith('_')} for r in coerced_rows]
    return ValidationResult(valid_rows=valid_rows, duplicates=duplicates, errors=errors)


def _fetch_fk_existing(fk_table: str, tenant_id: str, ids: list[str]) -> set[str]:
    if fk_table != 'dm_animals':
        return set(ids)
    try:
        from web_cabinet.insights_v1 import _conn
        with _conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT animal_id FROM dm_animals WHERE tenant_id=%s AND animal_id = ANY(%s)",
                    (tenant_id, ids),
                )
                return {r[0] for r in cur.fetchall()}
    except Exception as exc:
        logger.warning(f'_fetch_fk_existing failed: {exc}')
        return set()


def _fetch_existing_keys(spec: UploadType, tenant_id: str, rows: list[dict]) -> set[tuple]:
    if not rows or not spec.unique_key:
        return set()
    cols_to_select = [c for c in spec.unique_key if c != 'tenant_id']
    if not cols_to_select:
        return set()
    try:
        from web_cabinet.insights_v1 import _conn
        sql = f"SELECT {', '.join(cols_to_select)} FROM {spec.target_table} WHERE tenant_id=%s"
        with _conn() as conn:
            with conn.cursor() as cur:
                cur.execute(sql, (tenant_id,))
                existing = cur.fetchall()
        result: set[tuple] = set()
        for tup in existing:
            key = []
            for k in spec.unique_key:
                if k == 'tenant_id':
                    key.append(tenant_id)
                else:
                    idx = cols_to_select.index(k)
                    val = tup[idx]
                    if hasattr(val, 'isoformat'):
                        val = val.isoformat()[:10]
                    key.append(val)
            result.add(tuple(key))
        return result
    except Exception as exc:
        logger.warning(f'_fetch_existing_keys failed: {exc}')
        return set()


class TokenExpired(Exception):
    """Raised when preview_token is missing or expired or already consumed."""


class TenantMismatch(Exception):
    """Raised when commit caller's tenant_id does not match token's tenant_id."""


@dataclass
class _CacheEntry:
    type_id: str
    tenant_id: str
    valid_rows: list[dict]
    created_at: float


_TOKEN_CACHE: dict[str, _CacheEntry] = {}
_TOKEN_TTL_SECONDS = 5 * 60


def _gc_cache() -> None:
    now = time.time()
    expired = [k for k, v in _TOKEN_CACHE.items() if now - v.created_at > _TOKEN_TTL_SECONDS]
    for k in expired:
        _TOKEN_CACHE.pop(k, None)


def store_preview_token(type_id: str, tenant_id: str, valid_rows: list[dict]) -> str:
    _gc_cache()
    token = secrets.token_hex(8)
    _TOKEN_CACHE[token] = _CacheEntry(
        type_id=type_id, tenant_id=tenant_id,
        valid_rows=list(valid_rows), created_at=time.time(),
    )
    return token


def commit_rows(token: str, *, tenant_id: str, farm_id: Optional[str] = None,
                site_id: Optional[str] = None) -> UploadCommitResponse:
    _gc_cache()
    entry = _TOKEN_CACHE.pop(token, None)
    if entry is None:
        raise TokenExpired(token)
    if entry.tenant_id != tenant_id:
        _TOKEN_CACHE[token] = entry  # restore — don't consume on auth failure
        raise TenantMismatch()
    spec = TYPE_REGISTRY[entry.type_id]
    inserted = _do_insert(spec, entry.valid_rows, tenant_id=tenant_id,
                          farm_id=farm_id, site_id=site_id)
    return UploadCommitResponse(inserted=inserted, skipped_duplicates=0)


def _do_insert(spec: UploadType, rows: list[dict], *, tenant_id: str,
               farm_id: Optional[str], site_id: Optional[str]) -> int:
    if not rows:
        return 0
    auto_value = {
        'tenant_id': tenant_id,
        'farm_id': farm_id or _resolve_default_farm(tenant_id),
        'site_id': site_id or _resolve_default_site(tenant_id),
    }
    inserted = 0
    from web_cabinet.insights_v1 import _conn
    with _conn() as conn:
        with conn.cursor() as cur:
            for r in rows:
                full: dict[str, Any] = {}
                for col_name, src in spec.auto_fields.items():
                    full[col_name] = auto_value.get(src)
                if spec.server_id_column and spec.server_id_column not in r:
                    full[spec.server_id_column] = f'{spec.type_id[:3]}_{_uuid.uuid4().hex[:10]}'
                for col in spec.columns:
                    if col.name in r and r[col.name] is not None:
                        full[col.name] = r[col.name]
                full.setdefault('created_at', datetime.utcnow().isoformat())
                cols = list(full.keys())
                placeholders = ', '.join(['%s'] * len(cols))
                sql = (
                    f"INSERT INTO {spec.target_table} ({', '.join(cols)}) "
                    f"VALUES ({placeholders})"
                )
                try:
                    cur.execute(sql, [full[c] for c in cols])
                    inserted += cur.rowcount
                except Exception as exc:
                    logger.warning(f'insert row failed: {exc}; row={r}')
        conn.commit()
    return inserted


def _resolve_default_farm(tenant_id: str) -> Optional[str]:
    try:
        from web_cabinet.insights_v1 import _conn
        with _conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT farm_id FROM dm_farms WHERE tenant_id=%s LIMIT 1",
                    (tenant_id,),
                )
                row = cur.fetchone()
                return row[0] if row else None
    except Exception:
        return None


def _resolve_default_site(tenant_id: str) -> Optional[str]:
    try:
        from web_cabinet.insights_v1 import _conn
        with _conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT site_id FROM dm_sites WHERE tenant_id=%s LIMIT 1",
                    (tenant_id,),
                )
                row = cur.fetchone()
                return row[0] if row else None
    except Exception:
        return None


def run_preview(type_id: str, file_bytes: bytes, filename: str,
                tenant_id: str) -> UploadPreviewResponse:
    """Top-level helper: parse → validate → cache token → return response."""
    if type_id not in TYPE_REGISTRY:
        return UploadPreviewResponse(
            type=type_id, total_rows=0, valid=0, duplicates=0,
            errors=[UploadRowError(row=0, message=f'Unknown type: {type_id}')],
            preview_token='',
        )
    try:
        rows = parse_file(type_id, file_bytes, filename)
    except ValueError as exc:
        return UploadPreviewResponse(
            type=type_id, total_rows=0, valid=0, duplicates=0,
            errors=[UploadRowError(row=0, message=str(exc))],
            preview_token='',
        )
    result = validate_rows(type_id, rows, tenant_id=tenant_id)
    token = ''
    if result.valid_rows:
        token = store_preview_token(type_id, tenant_id, result.valid_rows)
    return UploadPreviewResponse(
        type=type_id,
        total_rows=len(rows),
        valid=len(result.valid_rows),
        duplicates=result.duplicates,
        errors=result.errors[:50],
        preview_token=token,
        valid_rows_sample=result.valid_rows[:5],
    )
