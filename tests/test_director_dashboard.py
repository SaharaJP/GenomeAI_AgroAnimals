from __future__ import annotations

from datetime import datetime
from pathlib import Path

from genomeai.kpi_v2 import run_kpi
from genomeai.dashboard_director import DirectorSummaryInputs, export_director_summary


def _xlsx_sheetnames(path: Path) -> set[str]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        try:
            return set(wb.sheetnames)
        finally:
            wb.close()
    except Exception:
        return set()


def test_director_export_creates_artifacts(tmp_path: Path):
    artifacts = tmp_path / "artifacts"
    dv = "dv_test_dash"
    # first create KPI run on fixtures
    run_kpi(
        data_version=dv,
        asof_date="2025-01-05",
        input_dir=Path("data/fixtures/target_v2"),
        artifacts_root=artifacts,
        config_kpi=Path("configs/kpi/kpi_v2.yaml"),
        config_thresholds=Path("configs/kpi/kpi_thresholds_v2.yaml"),
        run_id="kpi_test",
    )

    inputs = DirectorSummaryInputs(
        data_version=dv,
        artifacts_dir=artifacts,
        input_dir=None,
        kpi_run_id="kpi_test",
        asof_date=datetime.strptime("2025-01-05", "%Y-%m-%d").date(),
    )
    run_root = export_director_summary(inputs=inputs, run_id="dash_test")
    out_dir = run_root / "dashboards" / "director_summary"
    assert (out_dir / "director_summary.xlsx").exists()
    assert (out_dir / "director_summary.png").exists(), "T10-02: PNG snapshot is required"
    assert (out_dir / "kpi_plan_fact.csv").exists()
    assert (out_dir / "kpi_top_deviations.csv").exists()
    # pdf may be absent if reportlab missing; xlsx is required
    assert (out_dir / "dashboard_summary.json").exists()
    assert (run_root / "run_manifest.json").exists()
    assert (run_root / "checksums.json").exists()

    sheets = _xlsx_sheetnames(out_dir / "director_summary.xlsx")
    # openpyxl might be missing in ultra-minimal envs, so do a soft assert
    if sheets:
        assert "plan_fact" in sheets


def test_director_export_writes_trends_and_exceptions_when_input_dir_present(tmp_path: Path):
    artifacts = tmp_path / "artifacts"
    dv = "dv_test_dash_trends"
    run_kpi(
        data_version=dv,
        asof_date="2025-01-05",
        input_dir=Path("data/fixtures/target_v2"),
        artifacts_root=artifacts,
        config_kpi=Path("configs/kpi/kpi_v2.yaml"),
        config_thresholds=Path("configs/kpi/kpi_thresholds_v2.yaml"),
        run_id="kpi_test",
    )

    inputs = DirectorSummaryInputs(
        data_version=dv,
        artifacts_dir=artifacts,
        input_dir=Path("data/fixtures/target_v2"),
        kpi_run_id="kpi_test",
        asof_date=datetime.strptime("2025-01-05", "%Y-%m-%d").date(),
    )
    run_root = export_director_summary(inputs=inputs, run_id="dash_test_trends")
    out_dir = run_root / "dashboards" / "director_summary"

    assert (out_dir / "milk_trend_90d.csv").exists()
    assert (out_dir / "milk_trend_windows.csv").exists()
    # Always written (may be empty): stable artifact for UI
    assert (out_dir / "milk_trend_exceptions.csv").exists()
