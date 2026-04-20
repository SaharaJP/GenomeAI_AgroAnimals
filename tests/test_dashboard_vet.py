from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd


def test_compute_withdrawal_windows_basic() -> None:
    from genomeai.dashboard_vet import compute_withdrawal_windows

    tr = pd.DataFrame(
        [
            {
                "treatment_id": "TR1",
                "animal_id": "A1",
                "start_date": "2025-03-10",
                "end_date": "2025-03-12",
                "treatment_type": "antibiotic",
            },
            {
                "treatment_id": "TR2",
                "animal_id": "A2",
                "start_date": "2025-03-01",
                "end_date": "",
                "treatment_type": "unknown_type",
            },
        ]
    )
    rules = {
        "default_withdrawal_days": 7,
        "treatment_types": {"antibiotic": {"withdrawal_days": 10}},
    }

    out = compute_withdrawal_windows(tr, asof_date=date(2025, 3, 15), rules=rules)
    assert out.loc[0, "withdrawal_end_date_calc"].isoformat() == "2025-03-22"  # 2025-03-12 + 10d
    assert out.loc[1, "withdrawal_end_date_calc"].isoformat() == "2025-03-08"  # 2025-03-01 + 7d
    assert bool(out.loc[0, "withdrawal_active_asof"]) is True


def test_export_vet_registries_creates_xlsx(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from genomeai.dashboard_vet import VetDashboardInputs, export_vet_registries

    # input canonical
    input_dir = tmp_path / "input"
    input_dir.mkdir()

    pd.DataFrame(
        [{"tenant_id": "default", "animal_id": "A1", "farm_id": "F1", "ear_tag": "E1"}]
    ).to_csv(input_dir / "dm_animals.csv", index=False)
    pd.DataFrame(
        [{"tenant_id": "default", "event_id": "HE1", "animal_id": "A1", "event_date": "2025-03-10", "event_type": "mastitis"}]
    ).to_csv(input_dir / "dm_health_events.csv", index=False)
    pd.DataFrame(
        [
            {
                "tenant_id": "default",
                "treatment_id": "TR1",
                "animal_id": "A1",
                "start_date": "2025-03-10",
                "end_date": "2025-03-10",
                "treatment_type": "vitamin",
                "reason_event_id": "HE1",
                "withdrawal_end_date": "",
            }
        ]
    ).to_csv(input_dir / "dm_treatments.csv", index=False)
    pd.DataFrame(
        [{"tenant_id": "default", "alert_id": "AL1", "farm_id": "F1", "alert_date": "2025-03-10", "severity": "warn", "alert_type": "health_risk", "entity_type": "animal", "entity_id": "A1", "message": "risk flag"}]
    ).to_csv(input_dir / "dm_alerts.csv", index=False)

    artifacts_dir = tmp_path / "artifacts"
    artifacts_dir.mkdir()

    inputs = VetDashboardInputs(
        data_version="dv_test",
        artifacts_dir=artifacts_dir,
        input_dir=input_dir,
        asof_date=date(2025, 3, 11),
        qc_run=None,
    )
    run_root, xlsx_path = export_vet_registries(inputs=inputs, run_id="dash_test")
    assert run_root.exists()
    assert xlsx_path.exists()

    wb = load_workbook(xlsx_path)
    expected = {
        "health_events",
        "treatments",
        "active_withdrawal_animals",
        "inspection_list",
        "alerts_health_risk",
        "qc_alerts_auto",
        "qc_issues",
    }
    assert expected.issubset(set(wb.sheetnames))
