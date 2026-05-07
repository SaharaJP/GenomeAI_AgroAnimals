"""Template generation per UploadType."""
from __future__ import annotations

import csv
import io
import pytest


def test_list_types():
    from web_cabinet.uploads_v1 import list_types
    items = list_types()
    type_ids = {t.type for t in items}
    assert type_ids == {'milkings', 'health_events', 'animals', 'feed_rations'}


def test_csv_template_milkings():
    from web_cabinet.uploads_v1 import generate_template
    body, content_type, filename = generate_template('milkings', 'csv')
    assert content_type == 'text/csv; charset=utf-8'
    assert filename == 'milkings_template.csv'
    text = body.decode('utf-8-sig')  # BOM stripped
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    headers = rows[0]
    assert 'animal_id' in headers
    assert 'date' in headers
    assert 'milk_kg' in headers
    # second row is example
    assert len(rows) >= 2


def test_xlsx_template_milkings_parses_back():
    from web_cabinet.uploads_v1 import generate_template
    from openpyxl import load_workbook
    body, content_type, filename = generate_template('milkings', 'xlsx')
    assert filename == 'milkings_template.xlsx'
    wb = load_workbook(io.BytesIO(body))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert 'animal_id' in headers
    assert 'date' in headers
    wb.close()


@pytest.mark.parametrize('type_id', ['milkings', 'health_events', 'animals', 'feed_rations'])
@pytest.mark.parametrize('fmt', ['csv', 'xlsx'])
def test_template_smoke(type_id, fmt):
    from web_cabinet.uploads_v1 import generate_template
    body, content_type, filename = generate_template(type_id, fmt)
    assert len(body) > 50, f'{type_id} {fmt} template empty'
    assert filename.endswith(f'.{fmt}')


def test_unknown_type_raises():
    from web_cabinet.uploads_v1 import generate_template
    with pytest.raises(ValueError, match='unknown type'):
        generate_template('xxx', 'csv')


def test_unknown_format_raises():
    from web_cabinet.uploads_v1 import generate_template
    with pytest.raises(ValueError, match='unknown format'):
        generate_template('milkings', 'pdf')
