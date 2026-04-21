from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd

from core.operational_report_builder import build_operational_report_snapshot, build_operational_report_table, export_operational_report
from streamlit_app.saved_views_state import apply_saved_view_state, extract_saved_view_state


def _seed_input_dir(tmp_path: Path) -> Path:
    input_dir = tmp_path / 'input'
    input_dir.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([
        {'tenant_id': 'default', 'animal_id': 'A1001', 'farm_id': 'F1', 'site_id': 'S1', 'current_pen_id': 'P1', 'current_pen_name': 'Fresh', 'sex': 'F', 'birth_date': '2022-01-01', 'breed': 'Holstein', 'status': 'active'},
        {'tenant_id': 'default', 'animal_id': 'A1002', 'farm_id': 'F1', 'site_id': 'S1', 'current_pen_id': 'P2', 'current_pen_name': 'Hospital', 'sex': 'F', 'birth_date': '2021-06-01', 'breed': 'Jersey', 'status': 'active'},
        {'tenant_id': 'default', 'animal_id': 'A1003', 'farm_id': 'F2', 'site_id': 'S2', 'current_pen_id': 'P3', 'current_pen_name': 'Dry', 'sex': 'F', 'birth_date': '2020-05-01', 'breed': 'Holstein', 'status': 'dry'},
    ]).to_csv(input_dir / 'dm_animals.csv', index=False)
    pd.DataFrame([
        {'tenant_id': 'default', 'pen_id': 'P1', 'site_id': 'S1', 'pen_name': 'Fresh', 'pen_type': 'fresh', 'capacity_head': 50},
        {'tenant_id': 'default', 'pen_id': 'P2', 'site_id': 'S1', 'pen_name': 'Hospital', 'pen_type': 'hospital', 'capacity_head': 20},
        {'tenant_id': 'default', 'pen_id': 'P3', 'site_id': 'S2', 'pen_name': 'Dry', 'pen_type': 'dry', 'capacity_head': 30},
    ]).to_csv(input_dir / 'dm_pens.csv', index=False)
    pd.DataFrame([
        {'tenant_id': 'default', 'move_id': 'PM1', 'animal_id': 'A1001', 'from_pen_id': '', 'to_pen_id': 'P1', 'move_date': '2026-03-01', 'reason': 'fresh'},
        {'tenant_id': 'default', 'move_id': 'PM2', 'animal_id': 'A1002', 'from_pen_id': '', 'to_pen_id': 'P2', 'move_date': '2026-03-02', 'reason': 'check'},
        {'tenant_id': 'default', 'move_id': 'PM3', 'animal_id': 'A1003', 'from_pen_id': '', 'to_pen_id': 'P3', 'move_date': '2026-03-03', 'reason': 'dry'},
    ]).to_csv(input_dir / 'dm_pen_moves.csv', index=False)
    pd.DataFrame([
        {'tenant_id': 'default', 'animal_id': 'A1001', 'lactation_id': 'L1', 'parity': 2, 'calving_date': '2026-01-10', 'scc_cells_ml': 110000},
        {'tenant_id': 'default', 'animal_id': 'A1002', 'lactation_id': 'L2', 'parity': 3, 'calving_date': '2025-12-11', 'scc_cells_ml': 320000},
    ]).to_csv(input_dir / 'dm_lactations.csv', index=False)
    pd.DataFrame([
        {'tenant_id': 'default', 'event_id': 'HE1', 'animal_id': 'A1002', 'event_date': '2026-03-12', 'event_type': 'mastitis', 'severity': 'high', 'notes': 'Detected by SCC rise'},
        {'tenant_id': 'default', 'event_id': 'HE2', 'animal_id': 'A1001', 'event_date': '2026-03-05', 'event_type': 'metritis', 'severity': 'medium', 'notes': 'Follow-up'},
    ]).to_csv(input_dir / 'dm_health_events.csv', index=False)
    pd.DataFrame([
        {'tenant_id': 'default', 'repro_event_id': 'RE1', 'animal_id': 'A1001', 'event_date': '2026-03-08', 'event_type': 'insemination', 'bull_id': 'B1', 'result': 'unknown', 'notes': 'AI'},
        {'tenant_id': 'default', 'repro_event_id': 'RE2', 'animal_id': 'A1003', 'event_date': '2026-02-01', 'event_type': 'preg_check', 'bull_id': '', 'result': 'pregnant', 'notes': 'Confirmed'},
    ]).to_csv(input_dir / 'dm_repro_events.csv', index=False)
    pd.DataFrame([
        {'tenant_id': 'default', 'treatment_id': 'TR1', 'animal_id': 'A1002', 'start_date': '2026-03-12', 'end_date': '2026-03-20', 'treatment_type': 'antibiotic', 'reason_event_id': 'HE1', 'withdrawal_end_date': '2026-03-20'},
    ]).to_csv(input_dir / 'dm_treatments.csv', index=False)
    return input_dir


def test_t24_02_builds_health_repro_and_milk_quality_reports(tmp_path: Path) -> None:
    input_dir = _seed_input_dir(tmp_path)
    health = build_operational_report_snapshot(
        input_dir=input_dir,
        asof_date=date(2026, 3, 16),
        role='Vet',
        report_type='health_attention',
        filters={'animal_id': 'A1002'},
        limit=50,
    )
    assert health['report_type'] == 'health_attention'
    assert health['returned_rows'] == 2
    assert any(str(r.get('event_family')) == 'treatment' for r in health['rows'])

    repro = build_operational_report_snapshot(
        input_dir=input_dir,
        asof_date=date(2026, 3, 16),
        role='Zootech',
        report_type='repro_attention',
        filters={'animal_id': 'A1001'},
        limit=50,
    )
    assert repro['returned_rows'] == 1
    assert repro['rows'][0]['event_family'] == 'reproduction'

    milk = build_operational_report_snapshot(
        input_dir=input_dir,
        asof_date=date(2026, 3, 16),
        role='Director',
        report_type='milk_quality_watchlist',
        limit=50,
    )
    table = build_operational_report_table(milk)
    assert 'latest_scc_cells_ml' in table.columns
    assert any(str(r.get('animal_id')) == 'A1002' and str(r.get('milk_quality_flag')) in {'high_scc', 'treatment_withdrawal'} for r in milk['rows'])
    assert any(str(r.get('metric')) == 'high_scc_ge_200000' for r in milk['summary_rows'])


def test_t24_02_exports_include_summary_and_formulas_sheet(tmp_path: Path) -> None:
    input_dir = _seed_input_dir(tmp_path)
    snapshot = build_operational_report_snapshot(
        input_dir=input_dir,
        asof_date=date(2026, 3, 16),
        role='Director',
        report_type='groups_overview',
        limit=50,
    )
    xlsx_bytes = export_operational_report(snapshot, fmt='xlsx')
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    assert set(wb.sheetnames) >= {'report', 'summary', 'formulas'}
    ws = wb['report']
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert 'pen_name' in header and 'utilization_pct' in header


def test_t24_02_saved_view_state_supports_operational_report_builder() -> None:
    sess = {
        'operational_report_builder.asof': date(2026, 4, 2),
        'operational_report_builder.data_version': 'dv_t24_02',
        'operational_report_builder.report_type': 'health_attention',
        'operational_report_builder.q': 'mastitis',
        'operational_report_builder.farm_id': 'F1',
        'operational_report_builder.site_id': 'S1',
        'operational_report_builder.pen_id': 'P2',
        'operational_report_builder.animal_id': 'A1002',
        'operational_report_builder.status': 'active',
        'operational_report_builder.event_family': 'health',
        'operational_report_builder.event_type': 'mastitis',
        'operational_report_builder.severity': 'high',
        'operational_report_builder.sort_by': 'event_date_ts',
        'operational_report_builder.sort_dir': 'desc',
        'operational_report_builder.selected_columns': ['event_date', 'animal_id', 7],
        'operational_report_builder.limit': 80,
        'operational_report_builder.selected_row_id': 'HE1',
        'operational_report_builder.scc_threshold': 250000,
    }
    state = extract_saved_view_state(page_key='operational_report_builder', session_state=sess)
    assert state['operational_report_builder.asof'] == '2026-04-02'
    assert state['operational_report_builder.selected_columns'] == ['event_date', 'animal_id', '7']
    restored: dict[str, object] = {}
    apply_saved_view_state(page_key='operational_report_builder', state=state, session_state=restored)
    assert restored['operational_report_builder.asof'] == date(2026, 4, 2)
    assert restored['operational_report_builder.selected_columns'] == ['event_date', 'animal_id', '7']
    assert restored['operational_report_builder.scc_threshold'] == 250000


def test_t24_02_streamlit_contracts_docs_and_saved_views_present() -> None:
    page = Path('streamlit_app/pages/55_Operational_Report_Builder.py').read_text(encoding='utf-8')
    helper = Path('streamlit_app/operational_report_builder.py').read_text(encoding='utf-8')
    core = Path('src/core/operational_report_builder.py').read_text(encoding='utf-8')
    state = Path('streamlit_app/saved_views_state.py').read_text(encoding='utf-8')
    saved_views_page = Path('streamlit_app/pages/17_Saved_Views_And_Favorites.py').read_text(encoding='utf-8')
    favorites = Path('streamlit_app/personalization.py').read_text(encoding='utf-8')
    config = Path('configs/ui/ia_v3.yaml').read_text(encoding='utf-8')
    docs = Path('docs/operational_report_builder.md').read_text(encoding='utf-8')
    assumptions = Path('docs/assumptions.md').read_text(encoding='utf-8')

    assert 'Operational report builder' in page
    assert 'Export CSV' in page and 'Export XLSX' in page and 'Pin report' in page
    assert 'build_operational_report_snapshot' in helper and 'REPORT_TYPES' in core
    assert 'operational_report_builder' in state
    assert 'operational_report_builder' in saved_views_page
    assert 'operational_report' in favorites
    assert 'pages/55_Operational_Report_Builder.py' in config
    assert 'favorites' in docs.lower() and 'saved views' in docs.lower() and 'formulas' in docs.lower()
    assert '## T24-02 — operational report builder' in assumptions
